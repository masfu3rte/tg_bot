from datetime import datetime
from typing import Optional, TYPE_CHECKING, Iterable, Sequence
from xml.sax.saxutils import escape
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from config import Config


def safe_username(username: Optional[str], user_id: Optional[int] = None) -> str:
    if username:
        return f"@{username}" if not username.startswith("@") else username
    return f"id{user_id}" if user_id else "Пользователь"


def work_chat_user(username: Optional[str], user_id: int) -> str:
    """Format a user tag together with their ID for internal work-chat messages."""
    return f"{safe_username(username, user_id)} (id {user_id})"


def _internal_chat_id(chat_id: int) -> str:
    s = str(chat_id)
    if s.startswith("-100"):
        return s[4:]
    if s.startswith("-"):
        return s[1:]
    return s


def build_request_link(cfg: "Config", request: dict) -> Optional[str]:
    msg_id = request.get("channel_message_id")
    if not msg_id:
        return None
    internal = _internal_chat_id(cfg.REQUESTS_PUBLIC_CHANNEL_ID)
    return f"https://t.me/c/{internal}/{msg_id}"


def build_direct_link(cfg: "Config", message_id: int) -> str:
    internal = _internal_chat_id(cfg.REQUESTS_PUBLIC_CHANNEL_ID)
    return f"https://t.me/c/{internal}/{message_id}"


def write_simple_xlsx(path: str, sheet_name: str, rows: Iterable[Iterable[str]]) -> None:
    sheet_xml_rows = []
    for row_idx, row in enumerate(rows, start=1):
        cells = []
        for col_idx, value in enumerate(row, start=1):
            col_letter = chr(ord("A") + col_idx - 1)
            cell_ref = f"{col_letter}{row_idx}"
            cell_text = escape(str(value))
            cells.append(
                f'<c r="{cell_ref}" t="inlineStr"><is><t>{cell_text}</t></is></c>'
            )
        sheet_xml_rows.append(f'<row r="{row_idx}">{"".join(cells)}</row>')

    sheet_data = "".join(sheet_xml_rows)
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{sheet_data}</sheetData>"
        "</worksheet>"
    )

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        "<sheets>"
        f'<sheet name="{escape(sheet_name)}" sheetId="1" r:id="rId1"/>'
        "</sheets>"
        "</workbook>"
    )

    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )

    rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )

    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", rels_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def write_report_xlsx(
    path: str,
    summary_rows: Sequence[Sequence[object]],
    deal_rows: Sequence[Sequence[object]],
) -> None:
    """Write the formatted, two-sheet deal report without changing the legacy writer."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"
    deals = workbook.create_sheet("Сделки")

    for worksheet, rows in ((summary, summary_rows), (deals, deal_rows)):
        for row in rows:
            worksheet.append(list(row))
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for column_cells in worksheet.columns:
            width = min(
                max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells) + 2,
                45,
            )
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    deals.freeze_panes = "A2"
    if deals.max_row:
        deals.auto_filter.ref = deals.dimensions

    date_columns = (2, 3, 4)
    money_columns = (9, 10, 11)
    for row in deals.iter_rows(min_row=2):
        for index in date_columns:
            if isinstance(row[index - 1].value, datetime):
                row[index - 1].number_format = "DD.MM.YYYY HH:MM"
        for index in money_columns:
            row[index - 1].number_format = '#,##0.00'

    for row in summary.iter_rows(min_row=2):
        if row[0].value in {
            "Общий оборот завершенных сделок",
            "Общая комиссия сервиса",
            "Сумма выплат продавцам",
            "Средний чек",
            "Средняя комиссия",
        }:
            row[1].number_format = '#,##0.00'
        elif row[0].value == "Процент успешно завершенных сделок":
            row[1].number_format = '0.00"%"'

    workbook.save(path)
